"""Patch annotation workbooks with synthetic names and addresses.

The exporter writes review workbooks with a consistent schema:

``address`` / ``name`` / ``address.{lang}`` / ``name.{lang}`` /
``name.{lang}.final`` / ``review_notes.{lang}``

This module keeps the structural columns intact and replaces only the
value-bearing text fields.
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from seatau.annotation import addresses
from seatau.localization.synthetic_gen import (
    SyntheticValuePools,
    generate_value_pools,
    load_value_pools,
    save_value_pools,
)

META_SHEETS = frozenset({"Annotation guideline"})
LOCALIZABLE_COLUMN_PREFIX = "name"

_PERSON_CONTEXT_HINTS = {
    "contacts",
    "contact",
    "customer",
    "customers",
    "initialization_actions",
    "initial_state",
    "name",
    "passenger",
    "passengers",
    "persona",
    "profile",
    "user",
    "users",
}

_ADDRESS_CONTEXT_HINTS = {
    "address",
    "billing",
    "contact",
    "contacts",
    "shipping",
    "user",
    "users",
}

_PROSE_CONTEXT_KEYS = {
    "description",
    "expression",
    "known_info",
    "notes",
    "persona",
    "reason_for_call",
    "summary",
    "task_instructions",
    "ticket",
}

_PROSE_NAME_PATTERNS = (
    re.compile(
        r"(?i:\b(?:you are|your name is|customer name:|passenger name:|name:)\s+)"
        r"(?P<value>[A-Z][A-Za-z'’\-]+(?:[ \t]+[A-Z][A-Za-z'’\-]+){1,3})",
    ),
    re.compile(
        r"(?i:\b(?:customer name|passenger name|name)\s*:\s*)"
        r"(?P<value>[A-Z][A-Za-z'’\-]+(?:[ \t]+[A-Z][A-Za-z'’\-]+){1,3})",
    ),
)
_PROSE_PHONE_PATTERN = re.compile(
    r"\bphone number[: ]+\s*(?P<value>\d{3}-\d{3}-\d{4})\b",
    flags=re.IGNORECASE,
)
_PROSE_EMAIL_PATTERN = re.compile(
    r"\b(?P<value>[\w.+-]+@[\w.-]+\.[A-Za-z]{2,})\b",
    flags=re.IGNORECASE,
)
_PROSE_ZIP_PATTERN = re.compile(
    r"\bzip code[: ]+\s*(?P<value>\d{5})\b", flags=re.IGNORECASE
)
_PROSE_ROUTE_PATTERN = re.compile(
    r"(?i:\bfrom\s+)(?P<origin>[A-Z][A-Za-z'’.\-]+(?:\s+[A-Z][A-Za-z'’.\-]+)*)"
    r"(?i:\s+to\s+)"
    r"(?P<destination>[A-Z][A-Za-z'’.\-]+(?:\s+[A-Z][A-Za-z'’.\-]+)*)"
)
_PROSE_ADDRESS_PATTERNS = (
    re.compile(
        r"(?i:\b(?:shipping|billing|mailing|street|current)?\s*address(?: is|:)?\s+)"
        r"(?P<value>[^.\n;]+)",
    ),
    re.compile(
        r"(?i:\b(?:live|lives|living|located|residing|reside|stay|staying) at\s+)"
        r"(?P<value>[^.\n;]+)",
    ),
)

_PROSE_LOCATION_PREFIXES: dict[str, tuple[str, ...]] = {
    "id": ("di", "ke", "dari"),
    "tl": ("sa", "mula", "patungo sa", "papuntang"),
    "th": ("ที่", "ใน", "จาก", "ถึง"),
    "vi": ("ở", "tại", "đến", "từ"),
    "zh": ("在", "到", "从", "去"),
}
_PROSE_COUNTRY_PREFIXES: dict[str, tuple[str, ...]] = {
    "id": ("di luar negeri di", "di luar negeri ke", "di luar negeri"),
    "tl": ("sa ibang bansa sa", "ibang bansa sa", "nasa ibang bansa sa"),
    "th": ("อยู่ต่างประเทศใน", "อยู่ต่างประเทศที่", "ต่างประเทศใน", "ต่างประเทศที่"),
    "vi": ("ở nước ngoài tại", "ở nước ngoài ở", "nước ngoài tại", "nước ngoài ở"),
    "zh": ("在国外的", "在国外", "国外的", "国外"),
}


def _capitalized_location_pattern(prefixes: tuple[str, ...]) -> re.Pattern[str]:
    prefix_pattern = "|".join(
        sorted((re.escape(prefix) for prefix in prefixes), key=len, reverse=True)
    )
    return re.compile(
        rf"(?i:(?:^|(?<=\s))(?:{prefix_pattern})\s+)"
        rf"(?P<value>[A-Z][\w'’.\-]+(?:\s+[A-Z][\w'’.\-]+){{0,4}})"
    )


def _country_pattern(prefixes: tuple[str, ...]) -> re.Pattern[str]:
    prefix_pattern = "|".join(
        sorted((re.escape(prefix) for prefix in prefixes), key=len, reverse=True)
    )
    return re.compile(
        rf"(?i:(?:^|(?<=\s))(?:{prefix_pattern})\s+)"
        rf"(?P<value>[A-Z][\w'’.\-]+(?:\s+[A-Z][\w'’.\-]+){{0,4}})"
    )


@dataclass(frozen=True)
class ValueCase:
    """One source string that should be replaced in the workbook."""

    sheet: str
    row_number: int
    address: str
    category: str
    source: str


@dataclass
class WorkbookPatchReport:
    """Summary of a workbook localization pass."""

    workbook: Path
    lang: str
    output_path: Path
    cases: list[ValueCase] = field(default_factory=list)
    replacements: dict[str, int] = field(default_factory=dict)
    changed_cells: int = 0
    sample_pairs: list[tuple[str, str]] = field(default_factory=list)

    def summary(self) -> str:
        lines = [
            f"Workbook: {self.workbook}",
            f"Output:   {self.output_path}",
            f"Language: {self.lang}",
            f"Cases:    {len(self.cases)}",
            f"Changed cells: {self.changed_cells}",
        ]
        if self.replacements:
            lines.append("Replacements:")
            for category, count in sorted(self.replacements.items()):
                lines.append(f"  {category}: {count}")
        if self.sample_pairs:
            lines.append("Samples:")
            for src, dst in self.sample_pairs[:8]:
                lines.append(f"  {src!r} -> {dst!r}")
        return "\n".join(lines)


def _tokens_from_address(address_value: str) -> list[str]:
    return addresses.parse(address_value).body.split("/")


def _wordish(value: str) -> bool:
    return bool(re.fullmatch(r"[\w\s.'’\-]+", value, flags=re.UNICODE))


def _looks_like_person_name(value: str) -> bool:
    if any(ch.isdigit() for ch in value):
        return False
    parts = [part for part in re.split(r"\s+", value.strip()) if part]
    if len(parts) < 2 or len(parts) > 4:
        return False
    if not _wordish(value):
        return False
    return True


def _looks_like_address(value: str) -> bool:
    return bool(re.search(r"\d", value) and any(ch.isalpha() for ch in value))


def _category_for_row(address_value: str, text_value: str) -> str | None:
    tokens = _tokens_from_address(address_value)
    if not tokens:
        return None
    last = tokens[-1]
    parent_tokens = set(tokens[:-1])

    if last in {"first_name", "last_name", "city", "state", "country"}:
        return last
    if last in {"zip", "postcode"}:
        return "postcodes"
    if last in {"email", "phone", "phone_number"} and isinstance(text_value, str):
        return last
    if last == "address1":
        return "street_addresses"
    if last == "address2":
        return "secondary_addresses"
    if (
        last == "address"
        and isinstance(text_value, str)
        and _looks_like_address(text_value)
    ):
        return "full_addresses"
    if last == "name" and parent_tokens.intersection(_PERSON_CONTEXT_HINTS):
        if _looks_like_person_name(text_value):
            return "full_names"
    if last == "name" and parent_tokens.intersection(_ADDRESS_CONTEXT_HINTS):
        if _looks_like_address(text_value):
            return "full_addresses"
    return None


def _prose_cases(address_value: str, text_value: str, *, lang: str) -> list[ValueCase]:
    tokens = _tokens_from_address(address_value)
    if not tokens or tokens[-1] not in _PROSE_CONTEXT_KEYS:
        return []

    cases: list[ValueCase] = []
    for pattern in _PROSE_NAME_PATTERNS:
        for match in pattern.finditer(text_value):
            value = match.group("value").strip()
            if value:
                cases.append(
                    ValueCase(
                        sheet="",
                        row_number=0,
                        address=address_value,
                        category="full_names",
                        source=value,
                    )
                )
    for pattern in _PROSE_ADDRESS_PATTERNS:
        for match in pattern.finditer(text_value):
            value = match.group("value").strip()
            if value:
                cases.append(
                    ValueCase(
                        sheet="",
                        row_number=0,
                        address=address_value,
                        category="full_addresses",
                        source=value,
                    )
                )
    for match in _PROSE_ROUTE_PATTERN.finditer(text_value):
        origin = match.group("origin").strip()
        destination = match.group("destination").strip()
        if origin:
            cases.append(
                ValueCase(
                    sheet="",
                    row_number=0,
                    address=address_value,
                    category="city",
                    source=origin,
                )
            )
        if destination:
            cases.append(
                ValueCase(
                    sheet="",
                    row_number=0,
                    address=address_value,
                    category="city",
                    source=destination,
                )
            )
    location_pattern = _capitalized_location_pattern(
        _PROSE_LOCATION_PREFIXES.get(lang, _PROSE_LOCATION_PREFIXES["vi"])
    )
    for match in location_pattern.finditer(text_value):
        value = match.group("value").strip()
        if value:
            cases.append(
                ValueCase(
                    sheet="",
                    row_number=0,
                    address=address_value,
                    category="city",
                    source=value,
                )
            )
    country_pattern = _country_pattern(
        _PROSE_COUNTRY_PREFIXES.get(lang, _PROSE_COUNTRY_PREFIXES["vi"])
    )
    for match in country_pattern.finditer(text_value):
        value = match.group("value").strip()
        if value:
            cases.append(
                ValueCase(
                    sheet="",
                    row_number=0,
                    address=address_value,
                    category="countries",
                    source=value,
                )
            )
    if match := _PROSE_PHONE_PATTERN.search(text_value):
        value = match.group("value").strip()
        cases.append(
            ValueCase(
                sheet="",
                row_number=0,
                address=address_value,
                category="phone_number",
                source=value,
            )
        )
    if match := _PROSE_EMAIL_PATTERN.search(text_value):
        value = match.group("value").strip()
        cases.append(
            ValueCase(
                sheet="",
                row_number=0,
                address=address_value,
                category="email",
                source=value,
            )
        )
    if match := _PROSE_ZIP_PATTERN.search(text_value):
        value = match.group("value").strip()
        cases.append(
            ValueCase(
                sheet="",
                row_number=0,
                address=address_value,
                category="postcodes",
                source=value,
            )
        )
    return cases


def _iter_name_columns(headers: Iterable[Any]) -> list[int]:
    cols: list[int] = []
    for index, header in enumerate(headers):
        if isinstance(header, str) and header.startswith(LOCALIZABLE_COLUMN_PREFIX):
            cols.append(index)
    return cols


def scan_workbook(workbook: Path, *, lang: str) -> dict[str, list[ValueCase]]:
    """Collect source strings that should be localized in a workbook."""

    xls = load_workbook(workbook, read_only=True, data_only=False)
    cases: dict[str, list[ValueCase]] = defaultdict(list)

    for sheet in xls.sheetnames:
        if sheet in META_SHEETS:
            continue
        ws = xls[sheet]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        name_columns = _iter_name_columns(headers)
        if not name_columns:
            continue
        address_index = headers.index("address") if "address" in headers else None
        if address_index is None:
            continue
        for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
            address_value = row[address_index].value
            if not isinstance(address_value, str):
                continue
            for col_index in name_columns:
                cell_value = row[col_index].value
                if not isinstance(cell_value, str) or not cell_value.strip():
                    continue
                category = _category_for_row(address_value, cell_value)
                if category is None:
                    prose_cases = _prose_cases(address_value, cell_value, lang=lang)
                    for prose_case in prose_cases:
                        cases[prose_case.category].append(
                            ValueCase(
                                sheet=sheet,
                                row_number=row_number,
                                address=prose_case.address,
                                category=prose_case.category,
                                source=prose_case.source,
                            )
                        )
                    continue
                cases[category].append(
                    ValueCase(
                        sheet=sheet,
                        row_number=row_number,
                        address=address_value,
                        category=category,
                        source=cell_value,
                    )
                )
    return cases


def _replacement_field_name(category: str) -> str:
    return {
        "full_names": "full_names",
        "first_name": "first_names",
        "last_name": "last_names",
        "full_addresses": "full_addresses",
        "street_addresses": "street_addresses",
        "secondary_addresses": "secondary_addresses",
        "city": "cities",
        "state": "states",
        "country": "countries",
        "countries": "countries",
        "postcodes": "postcodes",
        "email": "emails",
        "phone": "phone_numbers",
        "phone_number": "phone_numbers",
    }[category]


def _generate_pool(
    lang: str,
    *,
    needed: dict[str, int],
    seed: int | None,
    pool_path: Path | None,
) -> SyntheticValuePools:
    if pool_path is not None and pool_path.exists():
        return load_value_pools(pool_path)

    count = max([32, *needed.values()])
    pools = generate_value_pools(lang, count=count, seed=seed)
    if pool_path is not None:
        save_value_pools(pools, pool_path)
    return pools


def _next_value(pool: list[str], index: int, *, category: str, lang: str) -> str:
    if not pool:
        raise RuntimeError(f"empty synthetic value pool for {category!r} in {lang!r}")
    return pool[index % len(pool)]


def _build_replacement_map(
    cases: dict[str, list[ValueCase]],
    pools: SyntheticValuePools,
) -> dict[str, str]:
    mapping: dict[str, str] = {}
    used: dict[str, set[str]] = defaultdict(set)
    counters: dict[str, int] = defaultdict(int)

    for category, value_cases in sorted(cases.items()):
        pool_attr = _replacement_field_name(category)
        pool = getattr(pools, pool_attr)
        unique_sources = sorted(
            {case.source for case in value_cases},
            key=lambda value: (-len(value), value),
        )
        for source in unique_sources:
            if source in mapping:
                continue
            value = _next_value(
                pool,
                counters[category],
                category=category,
                lang=pools.language,
            )
            counters[category] += 1
            while value in used[category]:
                value = _next_value(
                    pool,
                    counters[category],
                    category=category,
                    lang=pools.language,
                )
                counters[category] += 1
            mapping[source] = value
            used[category].add(value)
    return mapping


def _replacement_patterns(
    replacements: dict[str, str],
) -> list[tuple[re.Pattern[str], str]]:
    patterns: list[tuple[re.Pattern[str], str]] = []
    for source, target in sorted(
        replacements.items(), key=lambda item: (-len(item[0]), item[0])
    ):
        if re.fullmatch(r"[\w]+", source, flags=re.UNICODE):
            pattern = re.compile(rf"(?<!\w){re.escape(source)}(?!\w)")
        else:
            pattern = re.compile(re.escape(source))
        patterns.append((pattern, target))
    return patterns


def _replace_text(value: str, patterns: list[tuple[re.Pattern[str], str]]) -> str:
    new_value = value
    for pattern, replacement in patterns:
        new_value = pattern.sub(replacement, new_value)
    return new_value


def patch_annotation_workbook(
    workbook: Path,
    *,
    lang: str,
    output_path: Path | None = None,
    catalog_path: Path | None = None,
    seed: int | None = None,
    save_catalog: bool = True,
) -> WorkbookPatchReport:
    """Replace localized names and addresses in an annotation workbook."""

    cases = scan_workbook(workbook, lang=lang)
    needed = {
        category: len({case.source for case in value_cases})
        for category, value_cases in cases.items()
    }
    pools = _generate_pool(
        lang,
        needed=needed,
        seed=seed,
        pool_path=catalog_path if save_catalog else None,
    )
    if catalog_path is not None and not catalog_path.exists() and not save_catalog:
        save_value_pools(pools, catalog_path)

    replacements = _build_replacement_map(cases, pools)
    patterns = _replacement_patterns(replacements)
    out_path = output_path or workbook

    wb = load_workbook(workbook)
    changed_cells = 0
    for sheet in wb.sheetnames:
        if sheet in META_SHEETS:
            continue
        ws = wb[sheet]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        name_columns = _iter_name_columns(headers)
        for row in ws.iter_rows(min_row=2):
            for col_index in name_columns:
                cell = row[col_index]
                if not isinstance(cell.value, str) or not cell.value.strip():
                    continue
                patched = _replace_text(cell.value, patterns)
                if patched != cell.value:
                    cell.value = patched
                    changed_cells += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    sample_pairs = sorted(
        replacements.items(), key=lambda item: (-len(item[0]), item[0])
    )
    report = WorkbookPatchReport(
        workbook=workbook,
        lang=lang,
        output_path=out_path,
        cases=[case for value_cases in cases.values() for case in value_cases],
        replacements={
            category: len({case.source for case in value_cases})
            for category, value_cases in cases.items()
        },
        changed_cells=changed_cells,
        sample_pairs=sample_pairs[:12],
    )
    return report
