"""Export multilingual annotation workbook for domain artifacts.

CLI: ``python -m seatau.annotation export ...``

Produces ``annotation_{lang}.xlsx`` with one sheet per artefact (markdown,
JSON, TOML, python tools, python schema). Reviewers fill the
``name.{lang}.final`` column; the importer writes reviewed translations back
into ``data/tau2/domains/{domain}/{lang}/``.

Symmetric with :mod:`seatau.annotation.import_reviewed`. All shared
helpers (address taxonomy, markdown split, manifest I/O) live in sibling
modules — this file is concerned only with workbook layout and Excel
formatting.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import paths as path_defs
from seatau.annotation import addresses, manifests
from seatau.annotation.markdown import Section
from seatau.annotation.markdown import split as split_markdown
from seatau.annotation.python_tools import extract_tool_docstrings
from seatau.constants import to_project_relative_path
from seatau.translation.config import (
    DB_FILE_NAMES,
    DOMAIN_SKIPPED_TASK_FILES,
    MARKDOWN_GLOBS,
    PYTHON_FILES,
    SKIPPED_TASK_FILES,
    TASK_FILE_GLOBS,
    TOOL_PYTHON_FILES,
)
from seatau.translation.extractors import extract_files
from seatau.translation.models import DomainFile

INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
MAX_SHEET_NAME_LEN = 31


@dataclass(frozen=True)
class Artifact:
    """One translatable file in a domain (input to the exporter)."""

    domain: str
    path: Path
    kind: str  # policy | tasks | db | tools | tool_returns | schema


def annotation_columns(lang: str) -> list[str]:
    """Column order for an artefact sheet."""
    return [
        "address",
        "name",
        f"address.{lang}",
        f"name.{lang}",
        f"name.{lang}.final",
        f"review_notes.{lang}",
    ]


# -- Discovery -----------------------------------------------------------------


def discover_artifacts(
    domains: list[str],
    data_domains_root: Path,
    src_domains_root: Path,
) -> list[Artifact]:
    """Walk a domain dir and yield every artefact the pipeline knows about."""
    artifacts: list[Artifact] = []
    for domain in domains:
        data_dir = data_domains_root / domain
        src_dir = src_domains_root / domain
        if not data_dir.exists():
            raise FileNotFoundError(f"domain data dir not found: {data_dir}")
        if not src_dir.exists():
            raise FileNotFoundError(f"domain src dir not found: {src_dir}")
        seen: set[Path] = set()

        for pattern in MARKDOWN_GLOBS:
            for path in sorted(data_dir.glob(pattern)):
                if path in seen:
                    continue
                seen.add(path)
                artifacts.append(Artifact(domain=domain, path=path, kind="policy"))

        for pattern in TASK_FILE_GLOBS:
            domain_skip = DOMAIN_SKIPPED_TASK_FILES.get(domain, set())
            for path in sorted(data_dir.glob(pattern)):
                if path.name in SKIPPED_TASK_FILES or path.name in domain_skip:
                    continue
                if path in seen:
                    continue
                seen.add(path)
                artifacts.append(Artifact(domain=domain, path=path, kind="tasks"))

        for filename in DB_FILE_NAMES:
            path = data_dir / filename
            if not path.exists() or path in seen:
                continue
            seen.add(path)
            artifacts.append(Artifact(domain=domain, path=path, kind="db"))

        for filename in PYTHON_FILES:
            path = src_dir / filename
            if not path.exists() or path in seen:
                continue
            seen.add(path)
            kind = "tools" if filename in TOOL_PYTHON_FILES else "schema"
            artifacts.append(Artifact(domain=domain, path=path, kind=kind))
            if filename == "tools.py" and _tool_return_segments(domain, path):
                artifacts.append(
                    Artifact(domain=domain, path=path, kind="tool_returns")
                )
    return artifacts


def _tool_return_segments(domain: str, path: Path) -> list[tuple[tuple[str, str], str]]:
    domain_file = DomainFile(
        domain=domain,
        path=path,
        relative_path=path,
        kind="python",
    )
    extraction = extract_files([domain_file])
    rows: list[tuple[tuple[str, str], str]] = []
    for segment in extraction.segments:
        if (
            segment.kind == "tool_returns"
            and isinstance(segment.address, tuple)
            and len(segment.address) == 2
        ):
            rows.append(((segment.address[0], segment.address[1]), segment.text))
    return rows


def _load_manifest_assets(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    assets = payload.get("assets")
    return assets if isinstance(assets, dict) else {}


def find_translated_path(
    artifact: Artifact,
    lang_dir: Path,
    manifest_assets: dict[str, dict[str, Any]],
) -> Path | None:
    """Resolve the translated counterpart of an English artefact, if any."""
    if not lang_dir.exists():
        return None
    if artifact.kind == "tool_returns":
        direct = lang_dir / "tool_returns.json"
        if direct.exists():
            return direct

    rel_suffix = f"domains/{artifact.domain}/{artifact.path.name}"
    for asset in manifest_assets.values():
        for source_file in asset.get("source_files", []) or []:
            if not isinstance(source_file, dict):
                continue
            source_path = source_file.get("path")
            if isinstance(source_path, str) and source_path.endswith(rel_suffix):
                output_name = asset.get("output_file")
                if isinstance(output_name, str):
                    candidate = lang_dir / output_name
                    if candidate.exists():
                        return candidate

    if artifact.kind == "tool_returns":
        expected_name = "tool_returns.json"
    elif artifact.kind in {"tools", "schema"}:
        expected_name = f"{artifact.path.stem}.json"
    else:
        expected_name = artifact.path.name
    direct = lang_dir / expected_name
    if direct.exists():
        return direct

    if artifact.kind == "policy":
        candidates = sorted(lang_dir.glob(f"{artifact.path.stem}*.md"))
        if candidates:
            return max(candidates, key=lambda p: p.stat().st_mtime)

    if artifact.kind == "db":
        for ext in (".json", ".toml"):
            candidate = lang_dir / f"{artifact.path.stem}{ext}"
            if candidate.exists():
                return candidate

    return None


# -- Row builders --------------------------------------------------------------


def _structured_rows(path: Path, domain: str) -> list[tuple[str, str]]:
    """Extract `(address_body, text)` pairs from a JSON/TOML/python file."""
    domain_file = DomainFile(
        domain=domain,
        path=path,
        relative_path=path,
        kind=addresses.file_kind(path),
    )
    extraction = extract_files([domain_file])
    return [
        (addresses.body_from_segment(s.address), s.text) for s in extraction.segments
    ]


def _row(
    artifact: Artifact,
    body: str,
    english_text: str,
    translated_path: Path | None,
    translated_body: str,
    translated_text: str,
    lang: str,
) -> dict[str, str]:
    return {
        "address": addresses.format(artifact.path.name, body),
        "name": english_text,
        f"address.{lang}": (
            addresses.format(translated_path.name, translated_body)
            if translated_path
            else ""
        ),
        f"name.{lang}": translated_text,
        f"name.{lang}.final": "",
        f"review_notes.{lang}": "",
    }


def _policy_rows(
    artifact: Artifact, translated_path: Path | None, lang: str
) -> list[dict[str, str]]:
    english_sections = split_markdown(artifact.path.read_text(encoding="utf-8"))
    translated_sections: list[Section] = []
    if translated_path is not None and translated_path.exists():
        translated_sections = split_markdown(
            translated_path.read_text(encoding="utf-8")
        )
    rows: list[dict[str, str]] = []
    for idx, section in enumerate(english_sections):
        translated = (
            translated_sections[idx] if idx < len(translated_sections) else None
        )
        rows.append(
            _row(
                artifact=artifact,
                body=section.section_id,
                english_text=section.text,
                translated_path=translated_path,
                translated_body=translated.section_id if translated else "",
                translated_text=translated.text if translated else "",
                lang=lang,
            )
        )
    return rows


def _structured_artifact_rows(
    artifact: Artifact, translated_path: Path | None, lang: str
) -> list[dict[str, str]]:
    english_rows = _structured_rows(artifact.path, artifact.domain)
    translated_lookup: dict[str, str] = {}
    if translated_path is not None and translated_path.exists():
        translated_lookup = dict(_structured_rows(translated_path, artifact.domain))
    rows: list[dict[str, str]] = []
    for body, text in english_rows:
        rows.append(
            _row(
                artifact=artifact,
                body=body,
                english_text=text,
                translated_path=translated_path,
                translated_body=body,
                translated_text=translated_lookup.get(body, ""),
                lang=lang,
            )
        )
    return rows


def _tools_rows(
    artifact: Artifact, translated_path: Path | None, lang: str
) -> list[dict[str, str]]:
    english_docs = extract_tool_docstrings(artifact.path)
    translated_docs: dict[str, str] = {}
    if translated_path is not None and translated_path.exists():
        translated_docs = json.loads(translated_path.read_text(encoding="utf-8"))
    rows: list[dict[str, str]] = []
    for tool_name in sorted(english_docs):
        rows.append(
            _row(
                artifact=artifact,
                body=tool_name,
                english_text=english_docs[tool_name],
                translated_path=translated_path,
                translated_body=tool_name,
                translated_text=str(translated_docs.get(tool_name, "")),
                lang=lang,
            )
        )
    return rows


def _tool_return_rows(
    artifact: Artifact, translated_path: Path | None, lang: str
) -> list[dict[str, str]]:
    translated_lookup: dict[tuple[str, str], str] = {}
    if translated_path is not None and translated_path.exists():
        payload = json.loads(translated_path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            for section in ("exact", "templates"):
                section_payload = payload.get(section)
                if not isinstance(section_payload, dict):
                    continue
                for key, value in section_payload.items():
                    if isinstance(key, str) and isinstance(value, dict):
                        localized = value.get("localized")
                        if isinstance(localized, str):
                            translated_lookup[(section, key)] = localized

    rows: list[dict[str, str]] = []
    for (section, key), text in _tool_return_segments(artifact.domain, artifact.path):
        body = f"{section}/{key}"
        rows.append(
            _row(
                artifact=artifact,
                body=body,
                english_text=text,
                translated_path=translated_path,
                translated_body=f"{section}/{key}/localized",
                translated_text=translated_lookup.get((section, key), ""),
                lang=lang,
            )
        )
    return rows


def build_rows_for_artifact(
    artifact: Artifact, translated_path: Path | None, lang: str
) -> list[dict[str, str]]:
    """Dispatch to the right row builder based on artefact kind."""
    if artifact.kind == "policy":
        return _policy_rows(artifact, translated_path, lang)
    if artifact.kind == "tools":
        return _tools_rows(artifact, translated_path, lang)
    if artifact.kind == "tool_returns":
        return _tool_return_rows(artifact, translated_path, lang)
    # tasks, db, schema all use the structured (tuple-path) extractor
    return _structured_artifact_rows(artifact, translated_path, lang)


def artifact_sheet_name(artifact: Artifact) -> str:
    """Per-sheet name convention: ``<domain>_<file stem>``."""
    if artifact.kind == "tool_returns":
        return f"{artifact.domain}_tool_returns"
    return f"{artifact.domain}_{artifact.path.stem}"


def sanitize_sheet_name(name: str, used: set[str]) -> str:
    """Excel-safe sheet name with dedup suffix."""
    cleaned = INVALID_SHEET_CHARS.sub("", str(name)).strip() or "Sheet"
    base = cleaned[:MAX_SHEET_NAME_LEN]
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{base[: MAX_SHEET_NAME_LEN - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


# -- Metadata sheets -----------------------------------------------------------


def build_guideline_rows(lang: str) -> list[dict[str, str]]:
    """Static reviewer guidance rendered into the first sheet."""
    return [
        {
            "section": "Workflow",
            "key": "edit_scope",
            "value": f"Annotators should update only name.{lang}.final and review_notes.{lang}.",
        },
        {
            "section": "Workflow",
            "key": "alignment",
            "value": f"Each row aligns English baseline (address/name) with machine "
            f"translation (address.{lang}/name.{lang}).",
        },
        {
            "section": "Policy",
            "key": "policy_split",
            "value": "Markdown files are split by heading. One row per heading block.",
        },
        {
            "section": "Terminology",
            "key": "canonical_tokens",
            "value": "Keep canonical IDs, function names, and schema keys unchanged "
            "(e.g. order_id, get_customer_by_id).",
        },
        {
            "section": "Field dictionary",
            "key": "address",
            "value": "Machine-readable location of the extracted value in the source artefact.",
        },
        {
            "section": "Field dictionary",
            "key": "name",
            "value": "English baseline extracted value.",
        },
        {
            "section": "Field dictionary",
            "key": f"address.{lang}",
            "value": f"Matched location in the {lang} artefact (if available).",
        },
        {
            "section": "Field dictionary",
            "key": f"name.{lang}",
            "value": f"Machine-translated value in {lang}.",
        },
        {
            "section": "Field dictionary",
            "key": f"name.{lang}.final",
            "value": f"Human-localized final value in {lang}.",
        },
        {
            "section": "Field dictionary",
            "key": f"review_notes.{lang}",
            "value": "Optional reviewer notes or rationale.",
        },
    ]


def build_examples_rows(
    sheet_rows: list[tuple[str, list[dict[str, str]]]],
    lang: str,
    limit: int = 20,
) -> list[dict[str, str]]:
    """Pull the first few rows from each sheet to seed an Examples tab."""
    examples: list[dict[str, str]] = []
    for sheet_name, rows in sheet_rows:
        for row in rows:
            examples.append(
                {
                    "sheet": sheet_name,
                    "address": row.get("address", ""),
                    "name": row.get("name", ""),
                    f"address.{lang}": row.get(f"address.{lang}", ""),
                    f"name.{lang}": row.get(f"name.{lang}", ""),
                    f"name.{lang}.final": "",
                    f"review_notes.{lang}": "",
                }
            )
            if len(examples) >= limit:
                return examples
    return examples


def _to_dataframe(rows: list[dict[str, str]], lang: str) -> pd.DataFrame:
    cols = annotation_columns(lang)
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)


def _format_worksheet(worksheet, *, wide_text_columns: set[str]) -> None:
    """Apply reviewer-friendly worksheet formatting."""
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_idx, column_cells in enumerate(worksheet.columns, start=1):
        header = str(column_cells[0].value or "")
        if header in wide_text_columns:
            width = 60
        elif header.startswith("review_notes") or header.endswith(".final"):
            width = 36
        elif header.startswith("address"):
            width = 30
        else:
            width = 18
        worksheet.column_dimensions[get_column_letter(column_idx)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# -- CLI -----------------------------------------------------------------------


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Export an annotation workbook with one sheet per artifact "
            "and language-suffixed columns (e.g. name.vi)."
        )
    )
    parser.add_argument("--domains", nargs="+", default=["retail", "telecom"])
    parser.add_argument("--lang-id", default="vi", help="language code, e.g. vi")
    parser.add_argument("--data-domains-root", default="data/tau2/domains")
    parser.add_argument("--src-domains-root", default="src/tau2/domains")
    parser.add_argument("-o", "--output", required=True)
    parser.add_argument("--reviewer", default="unknown")
    parser.add_argument("--round-id", default=None)
    parser.add_argument("--annotation-metadata-dir", default=None)
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("-v", "--verbose", action="store_true")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    output_path = Path(args.output)
    if output_path.exists() and not args.overwrite:
        print(
            f"[ERROR] output exists (use --overwrite): {output_path}", file=sys.stderr
        )
        return 1

    data_domains_root = Path(args.data_domains_root)
    src_domains_root = Path(args.src_domains_root)
    metadata_dir = (
        Path(args.annotation_metadata_dir) if args.annotation_metadata_dir else None
    )

    try:
        artifacts = discover_artifacts(
            domains=args.domains,
            data_domains_root=data_domains_root,
            src_domains_root=src_domains_root,
        )
    except Exception as exc:
        print(f"[ERROR] failed to discover artifacts: {exc}", file=sys.stderr)
        return 1

    used: set[str] = set()
    sheet_payloads: list[tuple[str, list[dict[str, str]]]] = []
    missing: list[str] = []

    for artifact in artifacts:
        lang_dir = data_domains_root / artifact.domain / args.lang_id
        manifest_assets = _load_manifest_assets(lang_dir / "translation_manifest.json")
        translated_path = find_translated_path(artifact, lang_dir, manifest_assets)
        if translated_path is None:
            missing.append(
                f"{artifact.domain}:{artifact.path.name} (missing {args.lang_id})"
            )
        rows = build_rows_for_artifact(artifact, translated_path, args.lang_id)
        sheet_name = sanitize_sheet_name(artifact_sheet_name(artifact), used)
        sheet_payloads.append((sheet_name, rows))
        if args.verbose:
            tname = translated_path.name if translated_path else "none"
            print(
                f"[SHEET] {sheet_name} rows={len(rows)} "
                f"source={artifact.path.name} translated={tname}"
            )

    guideline_sheet = sanitize_sheet_name("Annotation guideline", used)
    examples_sheet = sanitize_sheet_name("Examples", used)
    guideline_rows = build_guideline_rows(args.lang_id)
    examples_rows = build_examples_rows(sheet_payloads, args.lang_id)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            pd.DataFrame(guideline_rows).to_excel(
                writer, sheet_name=guideline_sheet, index=False
            )
            pd.DataFrame(examples_rows).to_excel(
                writer, sheet_name=examples_sheet, index=False
            )
            for sheet_name, rows in sheet_payloads:
                _to_dataframe(rows, args.lang_id).to_excel(
                    writer, sheet_name=sheet_name, index=False
                )
            wide_cols = {
                "value",
                "name",
                f"name.{args.lang_id}",
                f"name.{args.lang_id}.final",
                f"review_notes.{args.lang_id}",
            }
            for worksheet in writer.book.worksheets:
                _format_worksheet(worksheet, wide_text_columns=wide_cols)
    except Exception as exc:
        print(f"[ERROR] failed to write workbook: {exc}", file=sys.stderr)
        return 1

    print(f"[OK] wrote annotation workbook: {output_path}")

    manifest_path = manifests.write_annotation_manifest(
        workbook=output_path,
        reviewer=args.reviewer,
        round_id=args.round_id or output_path.stem,
        lang=args.lang_id,
        domains=args.domains,
        data_domains_root=data_domains_root,
        src_domains_root=src_domains_root,
        sheet_rows={name: len(rows) for name, rows in sheet_payloads},
        missing_translations=missing,
        manifest_dir=metadata_dir,
        language_registry_path=to_project_relative_path(
            path_defs.LANGUAGES_PATH
        ).as_posix(),
    )
    print(f"[OK] wrote annotation manifest: {manifest_path}")
    if missing:
        print(f"[WARN] missing translated artifacts for '{args.lang_id}':")
        for item in missing:
            print(f"  - {item}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
