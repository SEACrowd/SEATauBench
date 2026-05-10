#!.venv/bin/env python3
"""Export multilingual annotation workbook for domain artifacts."""

from __future__ import annotations

import argparse
import ast
import json
import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from translation.config import (
    DB_FILE_NAMES,
    DOMAIN_SKIPPED_TASK_FILES,
    MARKDOWN_GLOBS,
    PYTHON_FILES,
    SKIPPED_TASK_FILES,
    TASK_FILE_GLOBS,
    TOOL_PYTHON_FILES,
)
from translation.extractors import extract_files
from translation.models import DomainFile

INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
MAX_SHEET_NAME_LEN = 31


@dataclass(frozen=True)
class Artifact:
    domain: str
    path: Path
    kind: str  # policy | tasks | db | tools | schema


@dataclass(frozen=True)
class MarkdownSection:
    section_id: str
    heading_line: str
    text: str


def annotation_columns(lang_id: str) -> list[str]:
    return [
        "address",
        "name",
        f"address.{lang_id}",
        f"name.{lang_id}",
        f"name.{lang_id}.final",
        f"review_notes.{lang_id}",
    ]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Export an annotation workbook with one sheet per artifact "
            "and language-suffixed columns (e.g., name.vi)."
        )
    )
    parser.add_argument(
        "--domains",
        nargs="+",
        default=["retail", "telecom"],
        help="Domain names to include (default: retail telecom).",
    )
    parser.add_argument(
        "--lang-id",
        default="vi",
        help="Language id suffix for translated columns (default: vi).",
    )
    parser.add_argument(
        "--data-domains-root",
        default="data/tau2/domains",
        help="Root directory containing domain data folders.",
    )
    parser.add_argument(
        "--src-domains-root",
        default="src/tau2/domains",
        help="Root directory containing domain source folders.",
    )
    parser.add_argument(
        "-o",
        "--output",
        required=True,
        help="Output .xlsx workbook path (e.g., data/sea-tau/annotation/review_vi.xlsx).",
    )
    parser.add_argument(
        "--reviewer",
        default="unknown",
        help="Reviewer identifier recorded in annotation metadata.",
    )
    parser.add_argument(
        "--round-id",
        default=None,
        help="Localization round identifier. Defaults to workbook stem.",
    )
    parser.add_argument(
        "--annotation-metadata-dir",
        default="config/sea-tau/annotation",
        help="Directory where annotation metadata manifests are stored.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite output file if it exists.",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Print detailed logs.",
    )
    return parser.parse_args()


def sanitize_sheet_name(name: str, used_names: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("", str(name)).strip() or "Sheet"
    base = cleaned[:MAX_SHEET_NAME_LEN]
    candidate = base
    counter = 2
    while candidate in used_names:
        suffix = f"_{counter}"
        allowed = MAX_SHEET_NAME_LEN - len(suffix)
        candidate = f"{base[:allowed]}{suffix}"
        counter += 1
    used_names.add(candidate)
    return candidate


def discover_artifacts(
    domains: list[str],
    data_domains_root: Path,
    src_domains_root: Path,
) -> list[Artifact]:
    artifacts: list[Artifact] = []
    for domain in domains:
        data_dir = data_domains_root / domain
        src_dir = src_domains_root / domain
        if not data_dir.exists():
            raise FileNotFoundError(f"domain data dir not found: {data_dir}")
        if not src_dir.exists():
            raise FileNotFoundError(f"domain src dir not found: {src_dir}")

        seen_paths: set[Path] = set()

        for pattern in MARKDOWN_GLOBS:
            for path in sorted(data_dir.glob(pattern)):
                if path in seen_paths:
                    continue
                seen_paths.add(path)
                artifacts.append(Artifact(domain=domain, path=path, kind="policy"))

        for pattern in TASK_FILE_GLOBS:
            domain_skipped_task_files = DOMAIN_SKIPPED_TASK_FILES.get(domain, set())
            for path in sorted(data_dir.glob(pattern)):
                if path.name in SKIPPED_TASK_FILES:
                    continue
                if path.name in domain_skipped_task_files:
                    continue
                if path in seen_paths:
                    continue
                seen_paths.add(path)
                artifacts.append(Artifact(domain=domain, path=path, kind="tasks"))

        for filename in DB_FILE_NAMES:
            path = data_dir / filename
            if not path.exists() or path in seen_paths:
                continue
            seen_paths.add(path)
            artifacts.append(Artifact(domain=domain, path=path, kind="db"))

        for filename in PYTHON_FILES:
            path = src_dir / filename
            if not path.exists() or path in seen_paths:
                continue
            seen_paths.add(path)
            kind = "tools" if filename in TOOL_PYTHON_FILES else "schema"
            artifacts.append(Artifact(domain=domain, path=path, kind=kind))
    return artifacts


def load_manifest_assets(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    assets = payload.get("assets")
    if isinstance(assets, dict):
        return assets
    return {}


def find_translated_path(
    artifact: Artifact,
    lang_dir: Path,
    manifest_assets: dict[str, dict[str, Any]],
) -> Path | None:
    if not lang_dir.exists():
        return None

    source_rel_suffix = f"domains/{artifact.domain}/{artifact.path.name}"
    for asset in manifest_assets.values():
        source_files = asset.get("source_files", [])
        if not isinstance(source_files, list):
            continue
        for source_file in source_files:
            if not isinstance(source_file, dict):
                continue
            source_path = source_file.get("path")
            if not isinstance(source_path, str):
                continue
            if source_path.endswith(source_rel_suffix):
                output_name = asset.get("output_file")
                if isinstance(output_name, str):
                    candidate = lang_dir / output_name
                    if candidate.exists():
                        return candidate

    expected_name = (
        f"{artifact.path.stem}.json"
        if artifact.kind in {"tools", "schema"}
        else artifact.path.name
    )
    direct = lang_dir / expected_name
    if direct.exists():
        return direct

    if artifact.kind == "policy":
        fallback_candidates = sorted(lang_dir.glob(f"{artifact.path.stem}*.md"))
        if fallback_candidates:
            return max(fallback_candidates, key=lambda p: p.stat().st_mtime)

    if artifact.kind == "db":
        base = artifact.path.stem
        for alt_name in (f"{base}.json", f"{base}.toml"):
            candidate = lang_dir / alt_name
            if candidate.exists():
                return candidate

    return None


def file_kind_for_path(path: Path) -> str:
    if path.suffix == ".md":
        return "markdown"
    if path.suffix == ".toml":
        return "toml"
    if path.suffix == ".py":
        return "python"
    return "json"


def segment_address_to_str(address: Any) -> str:
    if isinstance(address, tuple):
        return "/".join(address)
    return str(address)


def extract_structured_rows(path: Path, domain: str) -> list[tuple[str, str]]:
    domain_file = DomainFile(
        domain=domain,
        path=path,
        relative_path=path,
        kind=file_kind_for_path(path),
    )
    extraction = extract_files([domain_file])
    rows: list[tuple[str, str]] = []
    for segment in extraction.segments:
        rows.append((segment_address_to_str(segment.address), segment.text))
    return rows


def slugify(text: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", text.strip().lower())
    return slug.strip("-") or "section"


def split_markdown_by_heading(text: str) -> list[MarkdownSection]:
    lines = text.splitlines()
    sections: list[tuple[str, list[str]]] = []
    current_heading = "Preamble"
    current_lines: list[str] = []

    for line in lines:
        heading_match = re.match(r"^(#{1,6})\s+(.*)$", line)
        if heading_match:
            if current_lines:
                sections.append((current_heading, current_lines))
            current_heading = line.strip()
            current_lines = [line]
            continue
        current_lines.append(line)

    if current_lines:
        sections.append((current_heading, current_lines))

    dedupe: dict[str, int] = {}
    output: list[MarkdownSection] = []
    for idx, (heading_line, section_lines) in enumerate(sections, start=1):
        heading_text = heading_line
        if heading_line == "Preamble":
            heading_text = "preamble"
        slug = slugify(heading_text)
        dedupe[slug] = dedupe.get(slug, 0) + 1
        section_id = f"{idx:03d}_{slug}_{dedupe[slug]}"
        output.append(
            MarkdownSection(
                section_id=section_id,
                heading_line=heading_line,
                text="\n".join(section_lines).strip(),
            )
        )
    return output


def extract_tool_docstrings(path: Path) -> dict[str, str]:
    source = path.read_text(encoding="utf-8")
    tree = ast.parse(source)
    docs: dict[str, str] = {}
    for node in ast.walk(tree):
        if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            continue
        if not node.name or not node.name[0].islower() or node.name.startswith("_"):
            continue
        docstring = ast.get_docstring(node, clean=False)
        if docstring and docstring.strip():
            docs[node.name] = docstring.strip()
    return docs


def load_json_file(path: Path | None) -> dict[str, Any]:
    if path is None or not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def build_policy_rows(
    artifact: Artifact,
    translated_path: Path | None,
    lang_id: str,
) -> list[dict[str, str]]:
    english_sections = split_markdown_by_heading(
        artifact.path.read_text(encoding="utf-8")
    )
    translated_sections: list[MarkdownSection] = []
    if translated_path is not None and translated_path.exists():
        translated_sections = split_markdown_by_heading(
            translated_path.read_text(encoding="utf-8")
        )

    rows: list[dict[str, str]] = []
    for idx, section in enumerate(english_sections):
        translated_section = (
            translated_sections[idx] if idx < len(translated_sections) else None
        )
        rows.append(
            {
                "address": f"{artifact.path.name}::{section.section_id}",
                "name": section.text,
                f"address.{lang_id}": (
                    f"{translated_path.name}::{translated_section.section_id}"
                    if translated_path and translated_section is not None
                    else ""
                ),
                f"name.{lang_id}": (
                    translated_section.text if translated_section is not None else ""
                ),
                f"name.{lang_id}.final": "",
                f"review_notes.{lang_id}": "",
            }
        )
    return rows


def build_tasks_or_db_rows(
    artifact: Artifact,
    translated_path: Path | None,
    lang_id: str,
) -> list[dict[str, str]]:
    english_rows = extract_structured_rows(artifact.path, artifact.domain)
    translated_lookup: dict[str, str] = {}
    if translated_path is not None and translated_path.exists():
        translated_rows = extract_structured_rows(translated_path, artifact.domain)
        translated_lookup = {address: text for address, text in translated_rows}

    rows: list[dict[str, str]] = []
    for address, text in english_rows:
        translated_text = translated_lookup.get(address, "")
        rows.append(
            {
                "address": f"{artifact.path.name}::{address}",
                "name": text,
                f"address.{lang_id}": (
                    f"{translated_path.name}::{address}" if translated_path else ""
                ),
                f"name.{lang_id}": translated_text,
                f"name.{lang_id}.final": "",
                f"review_notes.{lang_id}": "",
            }
        )
    return rows


def build_tools_rows(
    artifact: Artifact,
    translated_path: Path | None,
    lang_id: str,
) -> list[dict[str, str]]:
    english_docs = extract_tool_docstrings(artifact.path)
    translated_docs = load_json_file(translated_path)

    rows: list[dict[str, str]] = []
    for tool_name in sorted(english_docs):
        rows.append(
            {
                "address": f"{artifact.path.name}::{tool_name}",
                "name": english_docs[tool_name],
                f"address.{lang_id}": (
                    f"{translated_path.name}::{tool_name}" if translated_path else ""
                ),
                f"name.{lang_id}": str(translated_docs.get(tool_name, "")),
                f"name.{lang_id}.final": "",
                f"review_notes.{lang_id}": "",
            }
        )
    return rows


def build_rows_for_artifact(
    artifact: Artifact,
    translated_path: Path | None,
    lang_id: str,
) -> list[dict[str, str]]:
    if artifact.kind == "policy":
        return build_policy_rows(artifact, translated_path, lang_id)
    if artifact.kind in {"tasks", "db"}:
        return build_tasks_or_db_rows(artifact, translated_path, lang_id)
    if artifact.kind == "tools":
        return build_tools_rows(artifact, translated_path, lang_id)
    if artifact.kind == "schema":
        return build_tasks_or_db_rows(artifact, translated_path, lang_id)
    return []


def artifact_sheet_name(artifact: Artifact) -> str:
    return f"{artifact.domain}_{artifact.path.stem}"


def build_guideline_rows(lang_id: str) -> list[dict[str, str]]:
    return [
        {
            "section": "Workflow",
            "key": "edit_scope",
            "value": (
                f"Annotators should update only name.{lang_id}.final and "
                f"review_notes.{lang_id}. Do not edit source columns."
            ),
        },
        {
            "section": "Workflow",
            "key": "alignment",
            "value": (
                f"Each row aligns English baseline (address/name) with machine "
                f"translation (address.{lang_id}/name.{lang_id})."
            ),
        },
        {
            "section": "Policy",
            "key": "policy_split",
            "value": (
                "Policy files are split by heading only. One row per heading block "
                "(including heading + its body)."
            ),
        },
        {
            "section": "Terminology",
            "key": "canonical_tokens",
            "value": (
                "Keep canonical IDs, function names, and schema keys unchanged "
                "(e.g., order_id, get_customer_by_id)."
            ),
        },
        {
            "section": "Field dictionary",
            "key": "address",
            "value": "Machine-readable location of the extracted value in the source artifact.",
        },
        {
            "section": "Field dictionary",
            "key": "name",
            "value": "English baseline extracted value.",
        },
        {
            "section": "Field dictionary",
            "key": f"address.{lang_id}",
            "value": f"Matched location in the {lang_id} artifact (if available).",
        },
        {
            "section": "Field dictionary",
            "key": f"name.{lang_id}",
            "value": f"Machine-translated value in {lang_id}.",
        },
        {
            "section": "Field dictionary",
            "key": f"name.{lang_id}.final",
            "value": f"Human-localized final value in {lang_id}.",
        },
        {
            "section": "Field dictionary",
            "key": f"review_notes.{lang_id}",
            "value": "Optional reviewer notes or rationale.",
        },
    ]


def build_examples_rows(
    sheet_rows: list[tuple[str, list[dict[str, str]]]],
    lang_id: str,
    limit: int = 20,
) -> list[dict[str, str]]:
    examples: list[dict[str, str]] = []
    for sheet_name, rows in sheet_rows:
        for row in rows:
            examples.append(
                {
                    "sheet": sheet_name,
                    "address": row.get("address", ""),
                    "name": row.get("name", ""),
                    f"address.{lang_id}": row.get(f"address.{lang_id}", ""),
                    f"name.{lang_id}": row.get(f"name.{lang_id}", ""),
                    f"name.{lang_id}.final": "",
                    f"review_notes.{lang_id}": "",
                }
            )
            if len(examples) >= limit:
                return examples
    return examples


def to_dataframe(rows: list[dict[str, str]], lang_id: str) -> pd.DataFrame:
    columns = annotation_columns(lang_id)
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns)


def format_worksheet(worksheet, *, wide_text_columns: set[str] | None = None) -> None:
    """Apply reviewer-friendly worksheet formatting."""
    wide_text_columns = wide_text_columns or set()
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_idx, column_cells in enumerate(worksheet.columns, start=1):
        header = str(column_cells[0].value or "")
        width = 18
        if header in wide_text_columns:
            width = 60
        elif header.startswith("review_notes") or header.endswith(".final"):
            width = 36
        elif header.startswith("address"):
            width = 30
        worksheet.column_dimensions[get_column_letter(column_idx)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _current_git_commit() -> str | None:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            check=True,
            capture_output=True,
            text=True,
        )
        return result.stdout.strip()
    except Exception:  # noqa: BLE001
        return None


def write_annotation_manifest(
    output_path: Path,
    metadata_dir: Path,
    reviewer: str,
    round_id: str,
    lang_id: str,
    domains: list[str],
    data_domains_root: Path,
    src_domains_root: Path,
    sheet_payloads: list[tuple[str, list[dict[str, str]]]],
    missing_translation_files: list[str],
) -> Path:
    metadata_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = metadata_dir / f"{output_path.stem}.manifest.yaml"

    manifest = {
        "manifest_version": 1,
        "generated_at": datetime.now(UTC).isoformat(),
        "reviewer": {"id": reviewer},
        "localization_round": {"id": round_id},
        "language": {"code": lang_id},
        "domains": sorted(domains),
        "workbook": {
            "path": str(output_path),
            "sheet_count": len(sheet_payloads) + 2,  # + guideline + examples
        },
        "source": {
            "data_domains_root": str(data_domains_root),
            "src_domains_root": str(src_domains_root),
            "language_registry": "config/languages.json",
            "git_commit": _current_git_commit(),
        },
        "artifacts": {
            "sheet_rows": {
                sheet_name: len(rows) for sheet_name, rows in sheet_payloads
            },
            "missing_translations": missing_translation_files,
        },
    }

    manifest_path.write_text(
        yaml.safe_dump(manifest, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )
    return manifest_path


def main() -> int:
    args = parse_args()
    output_path = Path(args.output)
    if output_path.exists() and not args.overwrite:
        print(
            f"[ERROR] output exists (use --overwrite): {output_path}",
            file=sys.stderr,
        )
        return 1

    data_domains_root = Path(args.data_domains_root)
    src_domains_root = Path(args.src_domains_root)
    metadata_dir = Path(args.annotation_metadata_dir)

    try:
        artifacts = discover_artifacts(
            domains=args.domains,
            data_domains_root=data_domains_root,
            src_domains_root=src_domains_root,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"[ERROR] failed to discover artifacts: {exc}", file=sys.stderr)
        return 1

    used_sheet_names: set[str] = set()
    sheet_payloads: list[tuple[str, list[dict[str, str]]]] = []
    missing_translation_files: list[str] = []

    for artifact in artifacts:
        lang_dir = data_domains_root / artifact.domain / args.lang_id
        manifest_assets = load_manifest_assets(lang_dir / "translation_manifest.json")
        translated_path = find_translated_path(
            artifact=artifact,
            lang_dir=lang_dir,
            manifest_assets=manifest_assets,
        )
        if translated_path is None:
            missing_translation_files.append(
                f"{artifact.domain}:{artifact.path.name} (missing {args.lang_id})"
            )
        rows = build_rows_for_artifact(
            artifact=artifact,
            translated_path=translated_path,
            lang_id=args.lang_id,
        )
        sheet_name = sanitize_sheet_name(
            artifact_sheet_name(artifact), used_sheet_names
        )
        sheet_payloads.append((sheet_name, rows))
        if args.verbose:
            translation_repr = translated_path.name if translated_path else "none"
            print(
                f"[SHEET] {sheet_name} rows={len(rows)} "
                f"source={artifact.path.name} translated={translation_repr}"
            )

    guideline_sheet_name = sanitize_sheet_name("Annotation guideline", used_sheet_names)
    examples_sheet_name = sanitize_sheet_name("Examples", used_sheet_names)
    guideline_rows = build_guideline_rows(args.lang_id)
    examples_rows = build_examples_rows(sheet_payloads, args.lang_id)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            pd.DataFrame(guideline_rows).to_excel(
                writer, sheet_name=guideline_sheet_name, index=False
            )
            pd.DataFrame(examples_rows).to_excel(
                writer, sheet_name=examples_sheet_name, index=False
            )
            for sheet_name, rows in sheet_payloads:
                to_dataframe(rows, args.lang_id).to_excel(
                    writer, sheet_name=sheet_name, index=False
                )
            for worksheet in writer.book.worksheets:
                format_worksheet(
                    worksheet,
                    wide_text_columns={
                        "value",
                        "name",
                        f"name.{args.lang_id}",
                        f"name.{args.lang_id}.final",
                        f"review_notes.{args.lang_id}",
                    },
                )
    except ModuleNotFoundError as exc:
        if exc.name == "openpyxl":
            print(
                "[ERROR] openpyxl is required to write .xlsx files. "
                "Install it with: pip install openpyxl",
                file=sys.stderr,
            )
            return 1
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1
    except Exception as exc:  # noqa: BLE001
        print(f"[ERROR] failed to write workbook: {exc}", file=sys.stderr)
        return 1

    print(f"[OK] wrote annotation workbook: {output_path}")
    manifest_path = write_annotation_manifest(
        output_path=output_path,
        metadata_dir=metadata_dir,
        reviewer=args.reviewer,
        round_id=args.round_id or output_path.stem,
        lang_id=args.lang_id,
        domains=args.domains,
        data_domains_root=data_domains_root,
        src_domains_root=src_domains_root,
        sheet_payloads=sheet_payloads,
        missing_translation_files=missing_translation_files,
    )
    print(f"[OK] wrote annotation manifest: {manifest_path}")
    if missing_translation_files:
        print(f"[WARN] missing translated artifacts for '{args.lang_id}':")
        for item in missing_translation_files:
            print(f"  - {item}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
