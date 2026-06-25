from __future__ import annotations

import importlib
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from seatau.annotation import import_reviewed
from seatau.annotation.export import main as export_main

import_module = importlib.import_module("seatau.annotation.import_reviewed")


def test_import_reviewed_writes_back_to_translated_language_dir(
    monkeypatch,
    tmp_path: Path,
) -> None:
    out_root = tmp_path / "data" / "tau2" / "domains"
    src_root = tmp_path / "src" / "tau2" / "domains"
    retail_src = src_root / "retail"
    retail_src.mkdir(parents=True)
    (out_root / "retail" / "vi").mkdir(parents=True)
    (retail_src / "tools.py").write_text(
        '''
class RetailTools:
    @is_tool
    def lookup_order(self):
        """Look up an order."""
'''.lstrip(),
        encoding="utf-8",
    )

    monkeypatch.setattr(import_module, "OUT_ROOT", out_root)
    monkeypatch.setattr(import_module, "SRC_ROOT", src_root)

    workbook = tmp_path / "annotation_vi.xlsx"
    pd.DataFrame(
        [
            {
                "address": "tools.py::lookup_order",
                "name": "Look up an order.",
                "address.vi": "tools.json::lookup_order",
                "name.vi": "Tra cứu đơn hàng.",
                "name.vi.final": "Tra cứu đơn hàng đã duyệt.",
                "review_notes.vi": "",
            }
        ]
    ).to_excel(workbook, sheet_name="retail_tools", index=False)

    report = import_reviewed(
        workbook,
        lang="vi",
        require_canonical_tokens=False,
        require_manifest=False,
    )

    reviewed_path = out_root / "retail" / "vi" / "tools.json"
    assert not report.errors
    assert reviewed_path.exists()
    assert json.loads(reviewed_path.read_text(encoding="utf-8")) == {
        "lookup_order": "Tra cứu đơn hàng đã duyệt."
    }
    assert not (out_root / "retail" / "vi_loc").exists()


def test_annotation_export_includes_tool_return_review_sheet(tmp_path: Path) -> None:
    workbook = tmp_path / "annotation_id.xlsx"
    manifest_dir = tmp_path / "manifests"

    exit_code = export_main(
        [
            "--domains",
            "telecom",
            "--lang-id",
            "id",
            "-o",
            str(workbook),
            "--annotation-metadata-dir",
            str(manifest_dir),
        ]
    )

    assert exit_code == 0
    wb = load_workbook(workbook, read_only=True)
    assert "telecom_tool_returns" in wb.sheetnames
    ws = wb["telecom_tool_returns"]
    assert ws.max_row == 13
    assert [ws.cell(1, col).value for col in range(1, 7)] == [
        "address",
        "name",
        "address.id",
        "name.id",
        "name.id.final",
        "review_notes.id",
    ]
    first_row = [ws.cell(2, col).value for col in range(1, 5)]
    assert first_row[0] == "tools.py::exact/line_suspended"
    assert first_row[2] == "tool_returns.json::exact/line_suspended/localized"


def test_import_reviewed_writes_tool_returns_json(monkeypatch, tmp_path: Path) -> None:
    out_root = tmp_path / "data" / "tau2" / "domains"
    src_root = tmp_path / "src" / "tau2" / "domains"
    telecom_src = src_root / "telecom"
    telecom_src.mkdir(parents=True)
    (out_root / "telecom" / "id").mkdir(parents=True)
    (telecom_src / "tools.py").write_text(
        """
TOOL_RETURN_MESSAGES = {
    "line_resumed": "Line resumed successfully",
    "bill_set_paid": "Bill {bill_id} set to paid",
}
""".lstrip(),
        encoding="utf-8",
    )

    monkeypatch.setattr(import_module, "OUT_ROOT", out_root)
    monkeypatch.setattr(import_module, "SRC_ROOT", src_root)

    workbook = tmp_path / "annotation_id.xlsx"
    pd.DataFrame(
        [
            {
                "address": "tools.py::exact/line_resumed",
                "name": "Line resumed successfully",
                "address.id": "tool_returns.json::exact/line_resumed/localized",
                "name.id": "Nomor berhasil dilanjutkan",
                "name.id.final": "Nomor aktif kembali.",
                "review_notes.id": "",
            },
            {
                "address": "tools.py::templates/bill_set_paid",
                "name": "Bill {bill_id} set to paid",
                "address.id": "tool_returns.json::templates/bill_set_paid/localized",
                "name.id": "Tagihan {bill_id} diatur menjadi dibayar",
                "name.id.final": "Tagihan {bill_id} ditandai lunas.",
                "review_notes.id": "",
            },
        ]
    ).to_excel(workbook, sheet_name="telecom_tool_returns", index=False)

    report = import_reviewed(
        workbook,
        lang="id",
        require_canonical_tokens=False,
        require_manifest=False,
    )

    reviewed_path = out_root / "telecom" / "id" / "tool_returns.json"
    payload = json.loads(reviewed_path.read_text(encoding="utf-8"))
    assert not report.errors
    assert payload["exact"]["line_resumed"] == {
        "source": "Line resumed successfully",
        "localized": "Nomor aktif kembali.",
    }
    assert payload["templates"]["bill_set_paid"] == {
        "pattern": "^Bill\\ (?P<bill_id>.+)\\ set\\ to\\ paid$",
        "source": "Bill {bill_id} set to paid",
        "localized": "Tagihan {bill_id} ditandai lunas.",
    }
