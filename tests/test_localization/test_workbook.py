from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from seatau.localization.patch_sheet import patch_annotation_workbook, scan_workbook


def _write_sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "telecom_tasks"
    ws.append(
        [
            "address",
            "name",
            "address.vi",
            "name.vi",
            "name.vi.final",
            "review_notes.vi",
        ]
    )
    ws.append(
        [
            "tasks.json::0/user_scenario/instructions/known_info",
            "You are John Smith with phone number 555-123-2002. You are currently abroad in France.",
            "tasks.json::0/user_scenario/instructions/known_info",
            "Bạn là John Smith với số điện thoại 555-123-2002. Hiện tại bạn đang ở nước ngoài tại Pháp.",
            "",
            "",
        ]
    )
    ws.append(
        [
            "tasks.json::0/user_scenario/instructions/reason_for_call",
            "You live at 123 Maple Street, Springfield, CA 94016. Email: john.smith@example.com. You are traveling from Seattle to Portland.",
            "tasks.json::0/user_scenario/instructions/reason_for_call",
            "Bạn sống tại 123 Maple Street, Springfield, CA 94016. Email: john.smith@example.com. Bạn đang di chuyển từ Seattle đến Portland.",
            "",
            "",
        ]
    )
    ws.append(
        [
            "tasks.json::0/ticket",
            "Customer name: John Smith, phone number: 555-123-2002, current location: abroad in France.",
            "tasks.json::0/ticket",
            "Tên khách hàng: John Smith, số điện thoại: 555-123-2002, vị trí hiện tại: ở nước ngoài tại Pháp.",
            "",
            "",
        ]
    )
    ws.append(
        [
            "db.toml::plans/0/name",
            "Basic Plan",
            "db.toml::plans/0/name",
            "Gói Cơ bản",
            "",
            "",
        ]
    )
    wb.save(path)


def test_scan_and_patch_workbook(tmp_path: Path) -> None:
    workbook = tmp_path / "annotation_vi.xlsx"
    output = tmp_path / "annotation_vi.localized.xlsx"
    _write_sample_workbook(workbook)

    cases = scan_workbook(workbook, lang="vi")
    assert "full_names" in cases
    assert "phone_number" in cases
    assert "full_addresses" in cases
    assert "email" in cases
    assert "city" in cases
    assert "postcodes" not in cases

    report = patch_annotation_workbook(
        workbook,
        lang="vi",
        output_path=output,
        seed=42,
    )

    assert report.changed_cells > 0
    assert output.exists()

    patched = load_workbook(output, read_only=True)
    ws = patched["telecom_tasks"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    assert "John Smith" not in rows[0][1]
    assert "555-123-2002" not in rows[0][1]
    assert "John Smith" not in rows[0][3]
    assert "555-123-2002" not in rows[0][3]
    assert "Pháp" not in rows[0][3]
    assert "123 Maple Street" not in rows[2][1]
    assert "john.smith@example.com" not in rows[2][1]
    assert "Seattle" not in rows[2][1]
    assert "123 Maple Street" not in rows[2][3]
    assert "john.smith@example.com" not in rows[2][3]
    assert "Seattle" not in rows[2][3]
    assert rows[3][1] == "Basic Plan"
    assert rows[0][3] != rows[0][1]
